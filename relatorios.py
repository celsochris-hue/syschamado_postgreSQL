"""
Módulo responsável pela geração de relatórios em Excel (.xlsx) e PDF.
"""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def gerar_excel(chamados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Chamados"

    colunas = [
        "ID", "Título", "Solicitante", "E-mail", "Setor", "Prioridade",
        "Status", "Responsável", "Aberto em", "Atualizado em", "Descrição",
    ]
    ws.append(colunas)

    preenchimento = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    for celula in ws[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal="center", vertical="center")

    for c in chamados:
        ws.append([
            c.id,
            c.titulo,
            c.solicitante,
            c.email_solicitante or "",
            c.setor or "",
            c.prioridade,
            c.status,
            c.responsavel.nome if c.responsavel else "",
            c.data_abertura.strftime("%d/%m/%Y %H:%M"),
            c.data_atualizacao.strftime("%d/%m/%Y %H:%M"),
            c.descricao,
        ])

    larguras = [6, 32, 22, 26, 16, 12, 12, 20, 18, 18, 55]
    for i, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{ws.max_row}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# PDF - lista de chamados
# ---------------------------------------------------------------------------
def gerar_pdf_lista(chamados, status_filtro="Todos"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    estilos = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("Relatório de Chamados", estilos["Title"]))
    elementos.append(Paragraph(
        f"Filtro de status: <b>{status_filtro}</b> &nbsp;|&nbsp; "
        f"Total: <b>{len(chamados)}</b> &nbsp;|&nbsp; "
        f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        estilos["Normal"]
    ))
    elementos.append(Spacer(1, 0.5 * cm))

    dados = [["ID", "Título", "Solicitante", "Setor", "Prioridade", "Status", "Responsável", "Abertura"]]
    for c in chamados:
        dados.append([
            str(c.id),
            c.titulo[:38],
            c.solicitante,
            c.setor or "-",
            c.prioridade,
            c.status,
            c.responsavel.nome if c.responsavel else "-",
            c.data_abertura.strftime("%d/%m/%Y"),
        ])

    tabela = Table(dados, repeatRows=1, colWidths=[1.2*cm, 6.8*cm, 3.6*cm, 3*cm, 2.6*cm, 2.6*cm, 3.5*cm, 2.5*cm])
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f5f9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elementos.append(tabela)

    doc.build(elementos)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# PDF - detalhe de um único chamado
# ---------------------------------------------------------------------------
def gerar_pdf_chamado(chamado):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=2 * cm, bottomMargin=2 * cm,
        leftMargin=2 * cm, rightMargin=2 * cm,
    )
    estilos = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"Chamado #{chamado.id}", estilos["Title"]))
    elementos.append(Paragraph(chamado.titulo, estilos["Heading2"]))
    elementos.append(Spacer(1, 0.3 * cm))

    info = [
        ["Status", chamado.status],
        ["Responsável", chamado.responsavel.nome if chamado.responsavel else "Não atribuído"],
        ["Solicitante", chamado.solicitante],
        ["E-mail", chamado.email_solicitante or "-"],
        ["Setor", chamado.setor or "-"],
        ["Prioridade", chamado.prioridade],
        ["Aberto em", chamado.data_abertura.strftime("%d/%m/%Y %H:%M")],
        ["Última atualização", chamado.data_atualizacao.strftime("%d/%m/%Y %H:%M")],
    ]
    tabela_info = Table(info, colWidths=[4.5 * cm, 11.5 * cm])
    tabela_info.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#eef0fb")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elementos.append(tabela_info)
    elementos.append(Spacer(1, 0.5 * cm))

    elementos.append(Paragraph("Descrição", estilos["Heading3"]))
    elementos.append(Paragraph(chamado.descricao.replace("\n", "<br/>"), estilos["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    if chamado.anexos:
        elementos.append(Paragraph("Anexos", estilos["Heading3"]))
        for a in chamado.anexos:
            elementos.append(Paragraph(
                f"• {a.nome_original} — enviado em {a.data_upload.strftime('%d/%m/%Y %H:%M')}",
                estilos["Normal"]
            ))
        elementos.append(Spacer(1, 0.5 * cm))

    if chamado.historico:
        elementos.append(Paragraph("Histórico de Status", estilos["Heading3"]))
        hist_dados = [["Data", "Alteração", "Por", "Observação"]]
        for h in chamado.historico:
            alteracao = f"{h.status_anterior} → {h.status_novo}" if h.status_anterior else f"Criado ({h.status_novo})"
            autor = h.usuario.nome if h.usuario else "Sistema"
            hist_dados.append([
                h.data.strftime("%d/%m/%Y %H:%M"),
                alteracao,
                autor,
                h.observacao or "-",
            ])
        tabela_hist = Table(hist_dados, colWidths=[3 * cm, 4 * cm, 3.5 * cm, 5.5 * cm])
        tabela_hist.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        elementos.append(tabela_hist)

    doc.build(elementos)
    buffer.seek(0)
    return buffer
